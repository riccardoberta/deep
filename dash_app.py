from __future__ import annotations

import base64
import json
import os
import re
import shutil
import threading
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

import dash
import dash_bootstrap_components as dbc
from dash import Dash, Input, Output, State, dash_table, dcc, html, no_update

try:
    import plotly.graph_objects as _go
    _PLOTLY_AVAILABLE = True
except ImportError:
    _PLOTLY_AVAILABLE = False

try:  # pragma: no cover - cosmetic tweak
    from flask.cli import show_server_banner as _show_server_banner

    def _suppress_server_banner(*args, **kwargs) -> None:
        return None

    show_server_banner = _suppress_server_banner  # type: ignore
except Exception:  # pragma: no cover
    pass
from dotenv import load_dotenv
from openpyxl import load_workbook

from collaborations import CollaborationBuilder
from data_preparation import DataPreparation
from export import generate_member_pdf
from importer import Importer
from analyser import df_from_payloads, load_all_runs, query_llm
from thresholds import compute_scores, load_thresholds

_THRESHOLDS = load_thresholds()

load_dotenv()


def _env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y"}


@dataclass(frozen=True)
class AppSettings:
    input_folder: Path
    year_windows: str
    sleep_seconds: float
    fetch_scopus: bool
    fetch_unige: bool
    fetch_iris: bool
    data_dir: Path


def _load_settings() -> AppSettings:
    raw_sleep = os.getenv("SLEEP_SECONDS", "3.0")
    try:
        sleep_seconds = float(raw_sleep)
    except (TypeError, ValueError):
        sleep_seconds = 3.0

    input_folder = Path(os.getenv("INPUT_FOLDER", "./input")).expanduser()
    input_folder.mkdir(parents=True, exist_ok=True)
    data_dir = Path(os.getenv("DATA_DIR", "data")).expanduser()
    return AppSettings(
        input_folder=input_folder,
        year_windows=os.getenv("YEAR_WINDOWS", "15,10,5"),
        sleep_seconds=sleep_seconds,
        fetch_scopus=_env_bool("FETCH_SCOPUS", True),
        fetch_unige=_env_bool("FETCH_UNIGE", True),
        fetch_iris=_env_bool("FETCH_IRIS", True),
        data_dir=data_dir,
    )


ALLOWED_INPUT_SUFFIXES = {".xlsx", ".xlsm"}


def _list_input_files() -> List[Path]:
    folder = SETTINGS.input_folder
    folder.mkdir(parents=True, exist_ok=True)
    items = [
        path
        for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() in ALLOWED_INPUT_SUFFIXES
    ]
    return sorted(items, key=lambda p: p.name.lower())


def _input_file_options() -> List[Dict[str, str]]:
    return [{"label": path.name, "value": str(path.resolve())} for path in _list_input_files()]


def _default_input_file() -> Optional[str]:
    options = _input_file_options()
    return options[0]["value"] if options else None


def _safe_uploaded_path(filename: str) -> Path:
    sanitized = re.sub(r"[^A-Za-z0-9._-]", "_", Path(filename).name)
    if not sanitized:
        sanitized = "input.xlsx"
    suffix = Path(sanitized).suffix.lower() or ".xlsx"
    if suffix not in ALLOWED_INPUT_SUFFIXES:
        raise ValueError("Unsupported file type. Upload XLSX/XLSM workbooks.")
    base = Path(sanitized).stem or "input"
    counter = 1
    target = SETTINGS.input_folder / f"{base}{suffix}"
    while target.exists():
        target = SETTINGS.input_folder / f"{base}_{counter}{suffix}"
        counter += 1
    return target


def _save_uploaded_input_file(filename: str, contents: str) -> Path:
    _, data = contents.split(",", 1)
    decoded = base64.b64decode(data)
    target = _safe_uploaded_path(filename)
    target.write_bytes(decoded)
    return target


def _delete_input_file(path_value: Optional[str]) -> None:
    if not path_value:
        return
    path = Path(path_value)
    try:
        if not path.exists():
            return
        if path.resolve().parent != SETTINGS.input_folder.resolve():
            return
        path.unlink()
    except Exception:
        pass


def _build_input_preview(path_value: Optional[str], limit: Optional[int] = None) -> tuple[List[Dict[str, str]], List[Dict[str, Any]], str]:
    if not path_value:
        return [], [], "No workbook selected."
    path = Path(path_value)
    if not path.exists():
        return [], [], "Selected workbook was not found on disk."
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header: List[str] = []
        for row in rows:
            header = [str(value).strip() if value is not None else "" for value in row]
            if any(header):
                break
        if not header:
            workbook.close()
            return [], [], "Workbook is empty."
        data_rows: List[Dict[str, Any]] = []
        for row in rows:
            values = [row[idx] if idx < len(row) else "" for idx in range(len(header))]
            if not any(value is not None and str(value).strip() for value in values):
                continue
            record: Dict[str, Any] = {}
            for idx, column in enumerate(header):
                key = column or f"Column {idx + 1}"
                record[key] = values[idx]
            data_rows.append(record)
            if limit is not None and len(data_rows) >= limit:
                break
        workbook.close()
        columns: List[Dict[str, str]] = []
        seen_ids: set[str] = set()
        for idx, name in enumerate(header):
            display = name or f"Column {idx + 1}"
            col_id = re.sub(r"[^A-Za-z0-9_]+", "_", display) or f"column_{idx + 1}"
            if col_id in seen_ids:
                suffix = 1
                candidate = f"{col_id}_{suffix}"
                while candidate in seen_ids:
                    suffix += 1
                    candidate = f"{col_id}_{suffix}"
                col_id = candidate
            seen_ids.add(col_id)
            columns.append({"name": display, "id": col_id})

        normalized_data: List[Dict[str, Any]] = []
        for record in data_rows:
            normalized_record: Dict[str, Any] = {}
            for column in columns:
                display = column["name"]
                key = column["id"]
                normalized_record[key] = record.get(display, "")
            normalized_data.append(normalized_record)
        return columns, normalized_data, f"Previewing {path.name}"
    except Exception as exc:  # pragma: no cover - preview helper
        return [], [], f"Unable to read workbook: {exc}"


def _input_file_exists(path_value: Optional[str]) -> bool:
    return bool(path_value and Path(path_value).exists())


class ImportManager:
    """Run Importer in the background while keeping progress accessible to Dash callbacks."""

    def __init__(self, *, sleep_seconds: float, data_dir: Path) -> None:
        self.sleep_seconds = sleep_seconds
        self.data_dir = data_dir
        self._lock = threading.Lock()
        self._thread: Optional[threading.Thread] = None
        self._status: str = "idle"
        self._logs: List[str] = []
        self._error: Optional[str] = None
        self._result: Optional[Dict[str, Any]] = None
        self._started_at: Optional[str] = None
        self._finished_at: Optional[str] = None
        self._stop_event = threading.Event()

    def start(
        self,
        *,
        input_workbook: str,
        year_windows: Iterable[int],
        fetch_scopus: bool,
        fetch_unige: bool,
        fetch_iris: bool,
    ) -> None:
        with self._lock:
            if self._thread and self._thread.is_alive():
                raise RuntimeError("An import is already running.")
            self._status = "running"
            self._logs = []
            self._error = None
            self._result = None
            self._started_at = datetime.now(UTC).isoformat()
            self._finished_at = None
            self._stop_event = threading.Event()

        thread = threading.Thread(
            target=self._run_import,
            args=(input_workbook, tuple(year_windows), fetch_scopus, fetch_unige, fetch_iris),
            daemon=True,
        )
        self._thread = thread
        thread.start()

    def stop(self) -> None:
        with self._lock:
            if self._thread and self._thread.is_alive() and not self._stop_event.is_set():
                self._stop_event.set()
                timestamp = datetime.now(UTC).strftime("%H:%M:%S")
                self._logs.append(f"[{timestamp}] ⏹️ Stop requested by user.")

    def get_state(self) -> Dict[str, Any]:
        with self._lock:
            return {
                "status": self._status,
                "logs": list(self._logs),
                "error": self._error,
                "result": self._result,
                "started_at": self._started_at,
                "finished_at": self._finished_at,
            }

    def _run_import(
        self,
        input_workbook: str,
        year_windows: Tuple[int, ...],
        fetch_scopus: bool,
        fetch_unige: bool,
        fetch_iris: bool,
    ) -> None:
        def logger(message: str) -> None:
            timestamp = datetime.now(UTC).strftime("%H:%M:%S")
            with self._lock:
                self._logs.append(f"[{timestamp}] {message}")

        importer = Importer(
            input_workbook=input_workbook,
            year_windows=year_windows,
            sleep_seconds=self.sleep_seconds,
            fetch_scopus=fetch_scopus,
            fetch_unige=fetch_unige,
            fetch_iris=fetch_iris,
            data_dir=self.data_dir,
            logger=logger,
            should_stop=self._stop_event.is_set,
        )

        try:
            run_dir, payloads, metadata = importer.run()
            if metadata:
                _perform_elaborations(payloads, run_dir, metadata, logger)
            result = _build_run_store(run_dir, payloads, metadata)
            with self._lock:
                self._result = result
        except Exception as exc:  # pragma: no cover - surfaced in UI
            with self._lock:
                self._error = str(exc)
        finally:
            with self._lock:
                if self._stop_event.is_set() and not self._error:
                    self._status = "cancelled"
                else:
                    self._status = "failed" if self._error else "completed"
                self._finished_at = datetime.now(UTC).isoformat()


SETTINGS = _load_settings()
IMPORT_MANAGER = ImportManager(sleep_seconds=SETTINGS.sleep_seconds, data_dir=SETTINGS.data_dir)
DATA_PREPARER = DataPreparation()


def _perform_elaborations(
    payloads: List[Dict[str, Any]],
    run_dir: Path,
    metadata: Dict[str, Any],
    logger: Callable[[str], None],
) -> None:
    input_file = metadata.get("input_file") or (Path(_default_input_file()).name if _default_input_file() else None)
    outputs_written = False
    if input_file:
        input_path = next((path for path in SETTINGS.input_folder.glob("*") if path.name == input_file), None)
        workbook_path = str(input_path) if input_path else _default_input_file()
        try:
            summary_path = DATA_PREPARER.prepare(payloads, run_dir, workbook_path)
            logger(f"📘 Results workbook saved to {summary_path}")
            metadata["summary_path"] = str(summary_path)
            outputs_written = True
        except Exception as exc:  # pragma: no cover
            logger(f"⚠️ Results workbook failed: {exc}")
    else:
        logger("⚠️ Skipping results workbook: unable to determine input workbook.")

    windows = metadata.get("year_windows") or []
    try:
        builder = CollaborationBuilder(windows, logger=logger)
        builder.build(payloads, run_dir)
        logger("🔗 Collaboration graph generated.")
    except Exception as exc:  # pragma: no cover
        logger(f"⚠️ Collaboration graph failed: {exc}")

    if outputs_written:
        metadata["last_outputs_updated_at"] = datetime.now(UTC).isoformat()
        _write_metadata(run_dir, metadata)


def _parse_year_windows(value: str) -> List[int]:
    parts = [part.strip() for part in (value or "").split(",") if part.strip()]
    if not parts:
        raise ValueError("Year windows cannot be empty.")
    try:
        return [int(part) for part in parts]
    except ValueError as exc:
        raise ValueError("Year windows must contain integers separated by commas.") from exc


def _latest_run_dir(base_dir: Path) -> Optional[Path]:
    runs = _list_run_directories(base_dir)
    return runs[0] if runs else None


def _list_run_directories(base_dir: Path) -> List[Path]:
    if not base_dir.is_dir():
        return []
    pattern = re.compile(r"^(\d{4}_\d{2}_\d{2})_(\d+)$")
    candidates: List[Tuple[str, int, Path]] = []
    for child in base_dir.iterdir():
        if child.is_dir():
            match = pattern.match(child.name)
            if match:
                candidates.append((match.group(1), int(match.group(2)), child))
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return [item[2] for item in candidates]


def _load_payloads_from_dir(run_dir: Path) -> List[Dict[str, Any]]:
    source_dir = run_dir / "source"
    if not source_dir.is_dir():
        return []
    payloads: List[Dict[str, Any]] = []
    for path in sorted(source_dir.glob("*.json")):
        try:
            payloads.append(json.loads(path.read_text(encoding="utf-8")))
        except Exception:
            continue
    return payloads


def _load_metadata(run_dir: Path) -> Dict[str, Any]:
    path = run_dir / "metadata.json"
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _build_run_store(run_dir: Path | None, payloads: List[Dict[str, Any]], metadata: Dict[str, Any]) -> Dict[str, Any]:
    sorted_payloads = sorted(
        payloads,
        key=lambda item: (
            str(item.get("surname", "")).lower(),
            str(item.get("name", "")).lower(),
        ),
    )
    return {
        "run_dir": str(run_dir) if run_dir else None,
        "payloads": sorted_payloads,
        "metadata": metadata,
    }


def _history_file(run_value: Optional[str]) -> Optional[Path]:
    """Return the path to the persistent history JSON for a given run value."""
    if not run_value:
        return None
    if run_value == "__all__":
        return SETTINGS.data_dir / "_history_all.json"
    return Path(run_value) / "analysis_history.json"


def _load_history(run_value: Optional[str]) -> List[Dict[str, Any]]:
    hf = _history_file(run_value)
    if not hf or not hf.exists():
        return []
    try:
        return json.loads(hf.read_text(encoding="utf-8"))
    except Exception:
        return []


def _save_history(run_value: Optional[str], history: List[Dict[str, Any]]) -> None:
    hf = _history_file(run_value)
    if not hf:
        return
    try:
        hf.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def _load_run_store_for_value(value: Optional[str]) -> Dict[str, Any]:
    if not value:
        return {"run_dir": None, "payloads": [], "metadata": {}}
    path = Path(value)
    if not path.exists():
        return {"run_dir": None, "payloads": [], "metadata": {}}
    payloads = _load_payloads_from_dir(path)
    metadata = _load_metadata(path)
    return _build_run_store(path, payloads, metadata)


def _run_dropdown_options() -> List[Dict[str, str]]:
    options = []
    for path in _list_run_directories(SETTINGS.data_dir):
        meta_path = path / "metadata.json"
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
        except Exception:
            meta = {}
        input_file = Path(meta.get("input_file", "")).stem or path.name
        count = meta.get("source_count", "?")
        # path.name is like "2026_04_10_3" → "2026/04/10 #3"
        parts = path.name.split("_")
        if len(parts) == 4:
            date_label = f"{parts[0]}/{parts[1]}/{parts[2]} #{parts[3]}"
        else:
            date_label = path.name
        label = f"{date_label} – {input_file} ({count} membri)"
        options.append({"label": label, "value": str(path.resolve())})
    return options


def _sync_run_dropdown(preferred: Optional[str]) -> Tuple[List[Dict[str, str]], Optional[str]]:
    options = _run_dropdown_options()
    values = {option["value"] for option in options}
    if not options:
        return options, None
    if preferred not in values:
        preferred = options[0]["value"]
    return options, preferred


def _delete_run_directory(value: Optional[str]) -> bool:
    if not value:
        return False
    path = Path(value)
    try:
        if not path.exists():
            return False
        if path.resolve().parent != SETTINGS.data_dir.resolve():
            return False
        shutil.rmtree(path)
        return True
    except Exception:
        return False


def _regenerate_run_outputs(value: Optional[str]) -> str:
    if not value:
        return "⚠️ Select a run before regenerating outputs."
    run_dir = Path(value)
    if not run_dir.exists():
        return "⚠️ Run directory not found."
    payloads = _load_payloads_from_dir(run_dir)
    metadata = _load_metadata(run_dir)
    if not payloads:
        return "⚠️ Run has no payloads to process."

    messages: List[str] = []

    def _collector(message: str) -> None:
        messages.append(message)

    _perform_elaborations(payloads, run_dir, metadata or {}, _collector)
    return messages[-1] if messages else "✅ Outputs regenerated."


def _write_metadata(run_dir: Path, metadata: Dict[str, Any]) -> None:
    metadata_path = run_dir / "metadata.json"
    metadata_path.write_text(json.dumps(metadata, indent=2), encoding="utf-8")


_TREE_KEY_STYLE   = {"color": "#0550ae", "fontWeight": "500", "fontFamily": "monospace", "fontSize": "0.85rem"}
_TREE_INDENT      = {"paddingLeft": "1.1rem", "borderLeft": "2px solid #e9ecef", "marginLeft": "2px", "marginTop": "2px"}
_TREE_ITEM_STYLE  = {"marginBottom": "3px", "lineHeight": "1.5"}
_TREE_SUMMARY_STYLE = {"cursor": "pointer", "userSelect": "none", "marginBottom": "2px", "paddingLeft": "2px"}
# Dict keys that are collapsed by default even at level 1
_TREE_COLLAPSED_KEYS = {"teaching", "scores", "career", "location", "scopus_metrics", "iris_products"}


def _tree_leaf_value(value: Any) -> html.Span:
    if value is None:
        return html.Span("null", style={"color": "#6c757d", "fontFamily": "monospace", "fontSize": "0.85rem"})
    if isinstance(value, bool):
        return html.Span(str(value).lower(), style={"color": "#d63384", "fontFamily": "monospace", "fontSize": "0.85rem"})
    if isinstance(value, (int, float)):
        return html.Span(str(value), style={"color": "#b45309", "fontFamily": "monospace", "fontSize": "0.85rem"})
    text = str(value)
    display = text if len(text) <= 120 else f'{text[:120]}…'
    return html.Span(display, style={"color": "#198754", "fontFamily": "monospace", "fontSize": "0.85rem", "wordBreak": "break-all"})


def _build_json_tree(value: Any, label: str = "value", level: int = 0) -> html.Div | html.Details:
    # ── Dict ──────────────────────────────────────────────────────────────────
    if isinstance(value, dict):
        if not value:
            return html.Div([
                html.Span(label, style=_TREE_KEY_STYLE),
                html.Span(": {}", style={"color": "#6c757d", "fontFamily": "monospace", "fontSize": "0.85rem"}),
            ], style=_TREE_ITEM_STYLE)

        children = [_build_json_tree(val, str(key), level + 1) for key, val in value.items()]
        inner = html.Div(children, style=_TREE_INDENT)

        # Root level: skip wrapper, render children directly
        if level == 0:
            return html.Div(children)

        return html.Details(
            [
                html.Summary(html.Span(label, style=_TREE_KEY_STYLE), style=_TREE_SUMMARY_STYLE),
                inner,
            ],
            open=level <= 1 and label not in _TREE_COLLAPSED_KEYS,
            style=_TREE_ITEM_STYLE,
        )

    # ── List ──────────────────────────────────────────────────────────────────
    if isinstance(value, list):
        if not value:
            return html.Div([
                html.Span(label, style=_TREE_KEY_STYLE),
                html.Span(": []", style={"color": "#6c757d", "fontFamily": "monospace", "fontSize": "0.85rem"}),
            ], style=_TREE_ITEM_STYLE)

        children = [_build_json_tree(item, f"[{i}]", level + 1) for i, item in enumerate(value)]
        inner = html.Div(children, style=_TREE_INDENT)

        return html.Details(
            [
                html.Summary(html.Span(label, style=_TREE_KEY_STYLE), style=_TREE_SUMMARY_STYLE),
                inner,
            ],
            open=False,
            style=_TREE_ITEM_STYLE,
        )

    # ── Leaf ──────────────────────────────────────────────────────────────────
    return html.Div(
        [
            html.Span(label, style=_TREE_KEY_STYLE),
            html.Span(": ", style={"color": "#6c757d", "fontFamily": "monospace", "fontSize": "0.85rem"}),
            _tree_leaf_value(value),
        ],
        style=_TREE_ITEM_STYLE,
    )


def _radar_chart(payload: Dict[str, Any]) -> Optional[dcc.Graph]:
    if not _PLOTLY_AVAILABLE:
        return None
    scores = payload.get("scores") or {}
    a_score = (scores.get("articles")  or {}).get("score")
    c_score = (scores.get("citations") or {}).get("score")
    h_score = (scores.get("hindex")    or {}).get("score")
    if all(v is None for v in [a_score, c_score, h_score]):
        return None
    cats   = ["Articles", "Citations", "H-index", "Articles"]
    vals   = [a_score or 0, c_score or 0, h_score or 0, a_score or 0]
    fig = _go.Figure(_go.Scatterpolar(
        r=vals, theta=cats, fill="toself",
        fillcolor="rgba(13,110,253,0.15)",
        line=dict(color="#0d6efd", width=2),
    ))
    fig.update_layout(
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 1.4],
                            tickvals=[0.4, 0.8, 1.2], tickfont=dict(size=10)),
            angularaxis=dict(tickfont=dict(size=12)),
        ),
        showlegend=False,
        margin=dict(l=40, r=40, t=10, b=10),
        height=240,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    return dcc.Graph(figure=fig, config={"displayModeBar": False})


def _score_color(score: Optional[float]) -> str:
    if score is None:   return "secondary"
    if score >= 1.2:    return "success"
    if score >= 0.8:    return "primary"
    if score >= 0.4:    return "warning"
    return "danger"


def _ratio_color(ratio: Optional[float]) -> str:
    if ratio is None:   return "text-muted"
    if ratio >= 1.0:    return "text-success fw-semibold"
    if ratio >= 0.7:    return "text-warning fw-semibold"
    return "text-danger fw-semibold"


def _ratio_text(ratio: Optional[float]) -> str:
    return f"{ratio:.2f}" if ratio is not None else "N/D"


def _level_row(row_label: str, level: Dict[str, Any]) -> html.Tr:
    years     = level.get("years")
    value     = level.get("value")
    threshold = level.get("threshold")
    ratio     = level.get("ratio")

    label_text = f"{row_label} ({years}a)" if years else row_label

    if value is None or threshold is None:
        formula = html.Span("N/D", className="text-muted small")
    else:
        ratio_str = f"{ratio:.2f}" if ratio is not None else "—"
        formula = html.Span(
            [
                html.Span(f"{value}", className="fw-semibold"),
                html.Span(" / ", className="text-muted"),
                html.Span(f"{threshold}", className="fw-semibold"),
                html.Span(" = ", className="text-muted"),
                html.Span(ratio_str, className=f"fw-bold {_ratio_color(ratio)}"),
            ]
        )

    return html.Tr([
        html.Td(label_text, className="text-muted small pe-3", style={"whiteSpace": "nowrap"}),
        html.Td(formula, className="small text-end"),
    ])


def _indicator_card(label: str, block: Dict[str, Any]) -> dbc.Col:
    score = block.get("score")
    score_text = f"{score:.1f}" if score is not None else "N/D"
    color = _score_color(score)

    ratio_rows = [
        _level_row("Associate Prof.",  block.get("ii_fascia")   or {}),
        _level_row("Full Prof.",       block.get("i_fascia")    or {}),
        _level_row("Evaluator",     block.get("commissario") or {}),
    ]

    return dbc.Col(
        dbc.Card(
            [
                dbc.CardHeader(
                    dbc.Row(
                        [
                            dbc.Col(html.Span(label, className="fw-bold small"), className="align-self-center"),
                            dbc.Col(
                                dbc.Badge(score_text, color=color, className="float-end", style={"fontSize": "0.875rem"}),
                                className="text-end",
                            ),
                        ],
                        align="center",
                    ),
                    className="py-2 px-3",
                ),
                dbc.CardBody(
                    html.Table(
                        html.Tbody(ratio_rows),
                        className="w-100 mb-0",
                    ),
                    className="py-2 px-3",
                ),
            ],
            color=color,
            outline=True,
        ),
        md=4,
    )


def _member_detail_component(payload: Dict[str, Any]) -> html.Div:
    ssd = payload.get("ssd")
    metrics = payload.get("scopus_metrics", [])

    # Always recompute from live metrics so the display is never stale
    # even when the payload was generated with an older scores format.
    scores = compute_scores(ssd, metrics, _THRESHOLDS)

    scores_panel = dbc.Card(
        dbc.CardBody(
            [
                html.H6("Bibliometric Indicators", className="mb-2"),
                html.Div(
                    f"SSD: {ssd}" if ssd else "SSD non disponibile",
                    className="text-muted small mb-2",
                ),
                dbc.Row(
                    [
                        _indicator_card("Articles",  scores["articles"]),
                        _indicator_card("Citations", scores["citations"]),
                        _indicator_card("H-index",   scores["hindex"]),
                    ],
                    className="g-2",
                ),
                html.Div(
                    [
                        dbc.Badge("0.0 below Associate Prof.", color="danger",    className="me-1 mt-2"),
                        dbc.Badge("0.4 Associate Prof.",      color="warning",   className="me-1 mt-2"),
                        dbc.Badge("0.8 Full Prof.",           color="primary",   className="me-1 mt-2"),
                        dbc.Badge("1.2 Evaluator",         color="success",   className="me-1 mt-2"),
                        html.Span(" · ratio: ", className="text-muted small ms-1 me-1"),
                        html.Span("≥ 1.0", className="text-success small fw-semibold me-1"),
                        html.Span("threshold met,", className="text-muted small me-1"),
                        html.Span("0.7–1.0", className="text-warning small fw-semibold me-1"),
                        html.Span("close,", className="text-muted small me-1"),
                        html.Span("< 0.7", className="text-danger small fw-semibold me-1"),
                        html.Span("below", className="text-muted small"),
                    ],
                    className="mt-2",
                ),
            ]
        ),
        className="mb-3",
    )

    radar = _radar_chart(payload)
    radar_panel = dbc.Card(
        dbc.CardBody(radar),
        className="mb-3",
    ) if radar else None

    raw_panel = dbc.Card(
        dbc.CardBody(
            [
                html.Div(
                    _build_json_tree(payload),
                    style={"fontFamily": "monospace", "fontSize": "0.85rem", "lineHeight": "1.6"},
                ),
            ]
        ),
        className="mb-3",
    )

    panels = [scores_panel]
    if radar_panel:
        panels.append(radar_panel)
    panels.append(raw_panel)
    return html.Div(panels)

RUN_OPTIONS_INITIAL = _run_dropdown_options()
DEFAULT_RUN_SELECTION = RUN_OPTIONS_INITIAL[0]["value"] if RUN_OPTIONS_INITIAL else None
RUN_STORE_INITIAL = _load_run_store_for_value(DEFAULT_RUN_SELECTION)
DEFAULT_RUN_MESSAGE = "Select a run to explore or manage its outputs."
DEFAULT_INPUT_FILE = _default_input_file()
DEFAULT_PREVIEW_COLUMNS, DEFAULT_PREVIEW_DATA, DEFAULT_PREVIEW_MESSAGE = _build_input_preview(DEFAULT_INPUT_FILE)

app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
app.title = "DEEP"
app._favicon = "logo.png"

# ---------------------------------------------------------------------------
# Authentication – Flask-session-based login
# ---------------------------------------------------------------------------
_server = app.server
_server.secret_key = os.getenv("APP_SECRET_KEY", "dev-secret-key")
_APP_USERNAME = os.getenv("APP_USERNAME", "admin")
_APP_PASSWORD = os.getenv("APP_PASSWORD", "deep2024")

_LOGIN_HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DEEP – Login</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    body {{ background: #f0f2f5; }}
    .card {{ border: none; border-radius: 12px; }}
  </style>
</head>
<body>
<div class="d-flex justify-content-center align-items-center" style="min-height:100vh">
  <div class="card shadow-sm p-4" style="width:360px">
    <div class="text-center mb-4">
      <h4 class="fw-bold text-primary mb-0">DEEP</h4>
      <small class="text-muted">DITEN Evaluation and Evidence Platform</small>
    </div>
    {error}
    <form method="post" action="/login">
      <div class="mb-3">
        <label class="form-label small fw-semibold">Username</label>
        <input type="text" name="username" class="form-control" autofocus required>
      </div>
      <div class="mb-3">
        <label class="form-label small fw-semibold">Password</label>
        <input type="password" name="password" class="form-control" required>
      </div>
      <button type="submit" class="btn btn-primary w-100">Sign in</button>
    </form>
  </div>
</div>
</body>
</html>"""

_LOGIN_ERROR = '<div class="alert alert-danger py-2 small">Invalid username or password.</div>'

from flask import request as _freq, session as _fsession, redirect as _fredirect


@_server.route("/login", methods=["GET", "POST"])
def _login():
    if _freq.method == "POST":
        if (
            _freq.form.get("username") == _APP_USERNAME
            and _freq.form.get("password") == _APP_PASSWORD
        ):
            _fsession["authenticated"] = True
            return _fredirect(_freq.args.get("next") or "/")
        return _LOGIN_HTML.format(error=_LOGIN_ERROR), 401
    return _LOGIN_HTML.format(error=""), 200


@_server.route("/logout")
def _logout():
    _fsession.clear()
    return _fredirect("/login")


@_server.before_request
def _require_login():
    public = {"/login", "/logout"}
    if _freq.path in public or _freq.path.startswith("/_dash") or _freq.path.startswith("/assets"):
        return None
    if not _fsession.get("authenticated"):
        return _fredirect(f"/login?next={_freq.path}")


def _build_import_tab() -> dbc.Container:
    left_panel = dbc.Card(
        dbc.CardBody(
            [
                # ── Member list ──────────────────────────────────────────
                html.H5("Member list", className="mb-2"),
                dbc.Row(
                    [
                        dbc.Col(
                            dcc.Dropdown(
                                id="input-file-dropdown",
                                options=_input_file_options(),
                                value=DEFAULT_INPUT_FILE,
                                placeholder="Select a workbook from the input folder",
                                clearable=False,
                            ),
                        ),
                        dbc.Col(
                            dcc.Upload(
                                id="upload-input-file",
                                children=dbc.Button("Upload", color="secondary", className="w-100"),
                                multiple=False,
                                accept=".xlsx,.xlsm",
                            ),
                            width="auto",
                        ),
                        dbc.Col(
                            dbc.Button(
                                "Remove",
                                id="delete-input-btn",
                                color="danger",
                                outline=True,
                                className="w-100",
                                disabled=DEFAULT_INPUT_FILE is None,
                            ),
                            width="auto",
                        ),
                    ],
                    className="g-2 align-items-center",
                ),
                html.Hr(className="my-3"),
                # ── Settings ─────────────────────────────────────────────
                html.H5("Settings", className="mb-2"),
                dbc.Row(
                    [
                        dbc.Col(dbc.Label("Time windows", className="mb-0"), width="auto", className="align-self-center"),
                        dbc.Col(dbc.Input(id="year-windows", type="text", value=SETTINGS.year_windows), md=3),
                        dbc.Col(dbc.Label("Data sources", className="mb-0"), width="auto", className="align-self-center ms-3"),
                        dbc.Col(
                            dbc.Checklist(
                                id="fetch-options",
                                options=[
                                    {"label": "UNIGE",  "value": "unige"},
                                    {"label": "IRIS",   "value": "iris"},
                                    {"label": "Scopus", "value": "scopus"},
                                ],
                                value=[
                                    opt for opt, flag in {
                                        "scopus": SETTINGS.fetch_scopus,
                                        "unige":  SETTINGS.fetch_unige,
                                        "iris":   SETTINGS.fetch_iris,
                                    }.items() if flag
                                ],
                                switch=True,
                                inline=True,
                            ),
                        ),
                    ],
                    className="g-2 align-items-center",
                ),
                dbc.Row(
                    [
                        dbc.Col(dbc.Button("Start Import", id="start-import", color="primary"), width="auto"),
                        dbc.Col(dbc.Button("Stop Import", id="stop-import", color="danger", outline=True, disabled=True), width="auto"),
                        dbc.Col(html.Div(id="import-status-text", className="text-muted fw-semibold align-self-center")),
                    ],
                    className="g-2 align-items-center mt-2",
                ),
            ]
        ),
        className="shadow-sm",
    )

    right_panel = dbc.Card(
        dbc.CardBody(
            [
                html.H5("Import Log", className="mb-2"),
                html.Pre(
                    id="import-log",
                    style={
                        "flex": "1",
                        "minHeight": "0",
                        "overflowY": "auto",
                        "backgroundColor": "#f8f9fa",
                        "border": "1px solid #dee2e6",
                        "borderRadius": "4px",
                        "padding": "10px 12px",
                        "fontSize": "0.68rem",
                        "fontFamily": "monospace",
                        "lineHeight": "1.6",
                        "color": "#212529",
                        "whiteSpace": "pre-wrap",
                        "wordBreak": "break-word",
                        "margin": "0",
                    },
                ),
            ],
            style={"height": "100%", "display": "flex", "flexDirection": "column"},
        ),
        className="shadow-sm h-100",
    )

    preview_panel = dbc.Card(
        dbc.CardBody(
            [
                html.H5("Preview", className="mb-2"),
                dbc.Alert("", id="input-preview-message", color="light", className="mb-2 py-1 small"),
                dash_table.DataTable(
                    id="input-preview-table",
                    columns=DEFAULT_PREVIEW_COLUMNS,
                    data=DEFAULT_PREVIEW_DATA,
                    style_as_list_view=True,
                    style_table={"maxHeight": "340px", "overflowY": "auto", "overflowX": "auto"},
                    style_header={
                        "backgroundColor": "#f8f9fa",
                        "fontWeight": "600",
                        "fontSize": 13,
                        "color": "#495057",
                        "borderBottom": "2px solid #dee2e6",
                        "borderTop": "none",
                        "padding": "8px 10px",
                    },
                    style_cell={
                        "textAlign": "left",
                        "padding": "7px 10px",
                        "fontSize": 13,
                        "color": "#212529",
                        "borderBottom": "1px solid #f0f0f0",
                        "fontFamily": "inherit",
                    },
                    style_data_conditional=[
                        {"if": {"state": "active"}, "backgroundColor": "rgba(13,110,253,0.06)", "border": "none"},
                    ],
                    page_action="none",
                    cell_selectable=False,
                ),
            ]
        ),
        className="shadow-sm mt-3",
    )

    return dbc.Container(
        [
            dbc.Row(
                [
                    dbc.Col(left_panel,  md=7),
                    dbc.Col(right_panel, md=5),
                ],
                className="g-3 pt-3",
            ),
            preview_panel,
        ],
        fluid=True,
        className="px-3 pb-3",
    )


def _select_data_bar(dropdown_options: List[Dict[str, str]]) -> dbc.Card:
    return dbc.Card(
        dbc.CardBody(
            [
                dcc.Download(id="download-summary"),
                # Hidden elements kept for callback compatibility
                html.Div(id="current-run-label", style={"display": "none"}),
                html.Div(id="run-action-message", style={"display": "none"}),
                dbc.Row(
                    [
                        dbc.Col(
                            html.H5("Select data", className="mb-0"),
                            width="auto",
                            className="align-self-center",
                        ),
                        dbc.Col(
                            dcc.Dropdown(
                                id="run-dropdown",
                                options=dropdown_options,
                                value=dropdown_options[0]["value"] if dropdown_options else None,
                                placeholder="Select a run to explore",
                                clearable=False,
                            ),
                        ),
                        dbc.Col(
                            dbc.Button("Download", id="download-summary-btn", color="success", size="sm"),
                            width="auto",
                        ),
                        dbc.Col(
                            dbc.Button("Rebuild", id="regen-run-btn", color="secondary", size="sm"),
                            width="auto",
                        ),
                        dbc.Col(
                            dbc.Button("Delete", id="delete-run-btn", color="danger", outline=True, size="sm"),
                            width="auto",
                        ),
                    ],
                    align="center",
                    className="g-2",
                ),
            ],
            className="py-2",
        ),
        className="shadow-sm mb-0",
    )


# Base conditional styles for the member table (no row selected).
# Re-used by callbacks to reset or update the highlight.
_TABLE_STYLE_BASE: List[Dict[str, Any]] = [
    {"if": {"row_index": "odd"}, "backgroundColor": "#fafafa"},
    # Suppress the default active-cell blue border so only the row highlight is visible.
    {"if": {"state": "active"}, "backgroundColor": "rgba(0,0,0,0)", "border": "1px solid transparent"},
]


def _table_style_with_row(row_index: int) -> List[Dict[str, Any]]:
    return _TABLE_STYLE_BASE + [
        {"if": {"row_index": row_index}, "backgroundColor": "rgba(13,110,253,0.10)", "borderTop": "1px solid rgba(13,110,253,0.25)", "borderBottom": "1px solid rgba(13,110,253,0.25)"},
    ]


def _member_table_card() -> dbc.Card:
    return dbc.Card(
        dbc.CardBody(
            [
                html.H5("Members", className="mb-2"),
                dbc.Input(
                    id="member-search",
                    placeholder="Search by name or SSD…",
                    type="text",
                    size="sm",
                    className="mb-2",
                    debounce=False,
                ),
                html.Div(
                  dash_table.DataTable(
                    id="member-table",

                    columns=[
                        {"name": "",        "id": "inspect"},
                        {"name": "Surname", "id": "surname"},
                        {"name": "Name",    "id": "name"},
                        {"name": "SSD",     "id": "ssd"},
                        {"name": "_payload_idx", "id": "_payload_idx", "hidden": True},
                    ],
                    data=[],
                    style_as_list_view=True,
                    style_table={"overflowX": "auto", "minWidth": "100%"},
                    style_header={
                        "backgroundColor": "#f8f9fa",
                        "fontWeight": "600",
                        "fontSize": 13,
                        "color": "#495057",
                        "borderBottom": "2px solid #dee2e6",
                        "borderTop": "none",
                        "padding": "8px 10px",
                    },
                    style_cell={
                        "textAlign": "left",
                        "padding": "8px 10px",
                        "fontSize": 13,
                        "color": "#212529",
                        "borderBottom": "1px solid #f0f0f0",
                        "fontFamily": "inherit",
                        "cursor": "default",
                    },
                    style_cell_conditional=[
                        {"if": {"column_id": "inspect"}, "width": "36px", "textAlign": "center", "cursor": "pointer", "color": "#6c757d"},
                        {"if": {"column_id": "ssd"},     "maxWidth": "160px", "overflow": "hidden", "textOverflow": "ellipsis", "color": "#6c757d", "fontSize": 10},
                    ],
                    style_data_conditional=_TABLE_STYLE_BASE,
                    sort_action="native",
                    sort_mode="single",
                    page_action="none",
                    row_selectable=False,
                    cell_selectable=True,
                    tooltip_data=[],
                    tooltip_duration=None,
                  ),
                  style={"maxHeight": "930px", "overflowY": "auto", "overflowX": "hidden"},
                ),
            ],
        ),
        className="shadow-sm",
    )


def _member_detail_card() -> dbc.Card:
    return dbc.Card(
        dbc.CardBody(
            [
                dbc.Row(
                    [
                        dbc.Col(html.H5("Member details", className="mb-0"), className="align-self-center"),
                        dbc.Col(
                            dbc.Button("Download PDF", id="member-pdf-btn",
                                       color="secondary", size="sm", disabled=True),
                            width="auto",
                        ),
                    ],
                    className="g-2 mb-3 align-items-center",
                ),
                dcc.Download(id="member-pdf-download"),
                dcc.Loading(
                    html.Div("Select a member using the magnifier icon.", id="member-detail"),
                    id="member-detail-loading",
                    type="circle",
                    color="#0d6efd",
                    delay_show=0,
                ),
            ],
            style={"overflowY": "auto"},
        ),
        className="shadow-sm h-100",
    )


def _build_exploring_tab() -> dbc.Container:
    return dbc.Container(
        [
            dbc.Row(
                dbc.Col(_select_data_bar(_run_dropdown_options()), md=12),
                className="g-0 pt-3 pb-2",
            ),
            dbc.Row(
                [
                    dbc.Col(_member_table_card(),  md=4),
                    dbc.Col(_member_detail_card(), md=8),
                ],
                className="g-3 align-items-stretch",
            ),
            dbc.Row(
                dbc.Col(_build_comparison_card(), md=12),
                className="g-0 pb-3",
            ),
        ],
        fluid=True,
        className="px-3",
    )


_DEFAULT_OLLAMA_URL   = os.getenv("OLLAMA_URL",   "http://localhost:11434")
_DEFAULT_OLLAMA_MODEL = os.getenv("OLLAMA_MODEL",  "llama3")
_OLLAMA_STARTUP_WAIT  = 10  # seconds to wait for ollama to start


def _ensure_ollama() -> Optional[str]:
    """
    Check that Ollama is reachable.  If not, try to start it via `ollama serve`.
    Returns None on success, or an error string if Ollama cannot be reached.
    """
    import subprocess
    import time
    import urllib.request
    import urllib.error

    url = _DEFAULT_OLLAMA_URL.rstrip("/") + "/api/version"

    def _ping() -> bool:
        try:
            urllib.request.urlopen(url, timeout=3)
            return True
        except Exception:
            return False

    if _ping():
        return None  # already running

    # Not running — try to start it
    try:
        subprocess.Popen(
            ["ollama", "serve"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
    except FileNotFoundError:
        return "Ollama is not installed. Download it from https://ollama.com and try again."
    except Exception as exc:
        return f"Failed to start Ollama: {exc}"

    # Wait up to _OLLAMA_STARTUP_WAIT seconds for it to become responsive
    deadline = time.time() + _OLLAMA_STARTUP_WAIT
    while time.time() < deadline:
        time.sleep(1)
        if _ping():
            return None  # started successfully

    return (
        f"Ollama was launched but did not respond within {_OLLAMA_STARTUP_WAIT}s. "
        "Try running 'ollama serve' manually in a terminal."
    )


def _data_selector_card(
    dropdown_id: str,
    options: List[Dict[str, str]],
    include_all: bool = False,
) -> dbc.Card:
    all_options = ([{"label": "All runs", "value": "__all__"}] if include_all else []) + options
    default = "__all__" if include_all else (options[0]["value"] if options else None)
    return dbc.Card(
        dbc.CardBody(
            dbc.Row(
                [
                    dbc.Col(
                        html.H5("Select data", className="mb-0"),
                        width="auto",
                        className="align-self-center",
                    ),
                    dbc.Col(
                        dcc.Dropdown(
                            id=dropdown_id,
                            options=all_options,
                            value=default,
                            placeholder="Select a run",
                            clearable=False,
                        ),
                    ),
                ],
                align="center",
                className="g-2",
            ),
            className="py-2",
        ),
        className="shadow-sm mb-0",
    )


def _build_analysing_tab() -> dbc.Container:
    controls_card = dbc.Card(
        dbc.CardBody(
            [
                dbc.Row(
                    [
                        dbc.Col(html.H5("Ask a question", className="mb-0"), className="align-self-center"),
                        dbc.Col(
                            html.Span(
                                f"Model: {_DEFAULT_OLLAMA_MODEL}",
                                id="analysis-model-label",
                                className="text-muted small",
                            ),
                            width="auto",
                            className="align-self-center",
                        ),
                    ],
                    className="g-2 mb-3 align-items-center",
                ),
                # Hidden input keeps the model value accessible to the callback
                dbc.Input(id="analysis-model", type="hidden", value=_DEFAULT_OLLAMA_MODEL),
                dbc.Row(
                    [
                        dbc.Col(
                            dbc.Textarea(
                                id="analysis-question",
                                placeholder=(
                                    "e.g. Show all members with h-index above 30, sorted by h-index\n"
                                    "e.g. Show members whose article count in the last 5 years grew "
                                    "the most across imports"
                                ),
                                style={"resize": "none", "height": "100%", "minHeight": "90px"},
                            ),
                            md=10,
                            style={"display": "flex", "flexDirection": "column"},
                        ),
                        dbc.Col(
                            dbc.Button(
                                "Ask",
                                id="analysis-ask-btn",
                                color="primary",
                                style={"height": "100%", "width": "100%"},
                            ),
                            md=2,
                            style={"display": "flex"},
                        ),
                    ],
                    className="g-2",
                    style={"alignItems": "stretch"},
                ),
                html.Div(id="analysis-status", className="text-muted small mt-2"),
            ]
        ),
        className="shadow-sm mb-3",
    )

    result_card = dbc.Card(
        dbc.CardBody(
            [
                dbc.Row(
                    [
                        dbc.Col(html.H5("Results", className="mb-0"), className="align-self-center"),
                        dbc.Col(
                            dbc.Button("Download Excel", id="analysis-download-btn",
                                       color="success", size="sm", disabled=True),
                            width="auto",
                        ),
                    ],
                    className="g-2 mb-2 align-items-center",
                ),
                dcc.Loading(
                    html.Div(
                        "Ask a question above to see results here.",
                        id="analysis-result",
                        className="text-muted",
                    ),
                    id="analysis-result-loading",
                    type="circle",
                    color="#0d6efd",
                    delay_show=200,
                ),
                dcc.Download(id="analysis-download"),
            ]
        ),
        className="shadow-sm",
    )

    history_card = dbc.Card(
        dbc.CardBody(
            [
                dbc.Row(
                    [
                        dbc.Col(html.H5("History", className="mb-0"), className="align-self-center"),
                        dbc.Col(
                            dbc.Button("Clear", id="analysis-clear-history-btn",
                                       color="danger", outline=True, size="sm"),
                            width="auto",
                        ),
                    ],
                    className="g-2 mb-2 align-items-center",
                ),
                html.Div(id="analysis-history-panel", className="text-muted",
                         children="No questions asked yet."),
            ]
        ),
        className="shadow-sm mt-3",
    )

    _run_opts = _run_dropdown_options()
    return dbc.Container(
        [
            dbc.Row(
                dbc.Col(_data_selector_card("analysing-run-dropdown", _run_opts, include_all=True), md=12),
                className="g-0 pt-3 pb-2",
            ),
            dbc.Row(dbc.Col(controls_card,  md=12), className="g-0 pb-2"),
            dbc.Row(dbc.Col(result_card,    md=12), className="g-0"),
            dbc.Row(dbc.Col(history_card,   md=12), className="g-0"),
        ],
        fluid=True,
        className="px-3 pb-3",
    )



def _build_summary_tab() -> dbc.Container:
    _run_opts = _run_dropdown_options()
    return dbc.Container(
        [
            dbc.Row(
                dbc.Col(_data_selector_card("summary-run-dropdown", _run_opts), md=12),
                className="g-0 pt-3 pb-2",
            ),
            dbc.Row(dbc.Col(dbc.Card(
                dbc.CardBody(html.Div(id="summary-content", className="text-muted",
                                     children="Select a run above to see the department summary.")),
                className="shadow-sm",
            ), md=12), className="g-0"),
        ],
        fluid=True,
        className="px-3 pb-3",
    )


def _build_comparison_card() -> dbc.Card:
    return dbc.Card(
        dbc.CardBody(
            [
                html.H5("Compare members", className="mb-2"),
                dbc.Row(
                    [
                        dbc.Col(
                            dcc.Dropdown(
                                id="compare-members-dropdown",
                                options=[],
                                multi=True,
                                placeholder="Select members to compare…",
                                clearable=True,
                            ),
                        ),
                        dbc.Col(
                            dbc.Button("Compare", id="compare-btn", color="primary", size="sm"),
                            width="auto",
                            className="align-self-center",
                        ),
                    ],
                    className="g-2 align-items-center",
                ),
                html.Div(id="comparison-result", className="mt-3"),
                dbc.Button("Export Excel", id="comparison-export-btn", color="success",
                           size="sm", className="mt-2", disabled=True),
            ]
        ),
        className="shadow-sm mt-3",
    )


header = html.Div(
    dbc.Row(
        [
            dbc.Col(html.Img(src="/assets/logo.png", style={"height": "64px"}), width="auto"),
            dbc.Col(
                [
                    html.H1("DEEP", className="mb-1"),
                    html.Div("DITEN Evaluation and Evidence Platform", className="text-muted", style={"fontSize": "1.125rem"}),
                ],
                width="auto",
            ),
            dbc.Col(
                html.A("Logout", href="/logout", className="btn btn-outline-secondary btn-sm"),
                width="auto",
                className="ms-auto",
            ),
        ],
        align="center",
        className="g-3",
    ),
    className="mb-4",
)

app.layout = dbc.Container(
    [
        header,
        dbc.Card(
            dbc.CardBody(
                dcc.Tabs(
                    id="main-tabs",
                    value="tab-import",
                    children=[
                        dcc.Tab(label="Importing",  value="tab-import",     children=_build_import_tab()),
                        dcc.Tab(label="Exploring",  value="tab-exploring",  children=_build_exploring_tab()),
                        dcc.Tab(label="Analysing",  value="tab-analysing",  children=_build_analysing_tab()),
                        dcc.Tab(label="Summary",    value="tab-summary",    children=_build_summary_tab()),
                    ],
                )
            ),
            className="shadow-sm",
        ),
        dcc.Store(id="run-store", data=RUN_STORE_INITIAL),
        dcc.Store(id="selected-member-idx", data=None),
        dcc.Store(id="analysis-history", data=[]),
        dcc.Store(id="analysis-result-store", data=None),
        dcc.Store(id="ssd-export-store", data=None),
        dcc.Store(id="comparison-export-store", data=None),
        dcc.Download(id="ssd-export-download"),
        dcc.Download(id="comparison-export-download"),
        dcc.Interval(id="import-poll-interval", interval=2_000, disabled=True),
    ],
    fluid=True,
    className="py-4",
)


def _format_import_status(state: Dict[str, Any]) -> str:
    status = state.get("status")
    if status == "running":
        return "⏳ Import in progress..."
    if status == "completed":
        return "✅ Import completed."
    if status == "failed":
        return f"⚠️ Import failed: {state.get('error')}"
    if status == "cancelled":
        return "⏹️ Import cancelled."
    return "Ready."


def _format_run_meta(run_data: Dict[str, Any]) -> html.Div:
    metadata = run_data.get("metadata") or {}
    if not run_data.get("run_dir"):
        return html.Div("No run selected.", className="text-muted")
    items = []
    if metadata.get("year_windows"):
        items.append(html.Li(f"Year windows: {metadata['year_windows']}"))
    sources = [
        name
        for name, enabled in {
            "Scopus": metadata.get("fetch_scopus"),
            "UNIGE": metadata.get("fetch_unige"),
            "IRIS": metadata.get("fetch_iris"),
        }.items()
        if enabled
    ]
    if sources:
        items.append(html.Li("Sources: " + ", ".join(sources)))
    if metadata.get("created_at"):
        items.append(html.Li(f"Created at: {metadata['created_at']}"))
    if metadata.get("source_count") is not None:
        items.append(html.Li(f"Members: {metadata['source_count']}"))
    return html.Ul(items) if items else html.Div("No metadata available.", className="text-muted")


@app.callback(
    Output("input-file-dropdown", "options"),
    Output("input-file-dropdown", "value"),
    Output("input-preview-table", "columns"),
    Output("input-preview-table", "data"),
    Output("delete-input-btn", "disabled"),
    Output("input-preview-message", "children"),
    Output("upload-input-file", "contents"),
    Input("input-file-dropdown", "value"),
    Input("upload-input-file", "contents"),
    Input("delete-input-btn", "n_clicks"),
    State("upload-input-file", "filename"),
)
def manage_input_files(
    selected_value: Optional[str],
    upload_contents: Optional[str],
    delete_clicks: Optional[int],
    upload_filename: Optional[str],
):
    triggered = dash.ctx.triggered_id
    new_value = selected_value
    message_override: Optional[str] = None
    reset_upload = dash.no_update

    if triggered == "upload-input-file" and upload_contents and upload_filename:
        try:
            saved = _save_uploaded_input_file(upload_filename, upload_contents)
            new_value = str(saved.resolve())
            message_override = f"Uploaded {saved.name}"
        except ValueError as exc:
            message_override = f"⚠️ {exc}"
        except Exception as exc:  # pragma: no cover
            message_override = f"⚠️ Upload failed: {exc}"
        reset_upload = None
    elif triggered == "delete-input-btn" and delete_clicks and new_value:
        _delete_input_file(new_value)
        message_override = "Selected workbook deleted."
        new_value = None

    options = _input_file_options()
    option_values = {option["value"] for option in options}
    if not option_values:
        new_value = None
    elif new_value not in option_values:
        new_value = options[0]["value"]

    columns, data, preview_message = _build_input_preview(new_value)
    message = message_override or preview_message
    delete_disabled = new_value is None

    return options, new_value, columns, data, delete_disabled, message, reset_upload


@app.callback(
    Output("import-status-text", "children"),
    Output("import-log", "children"),
    Output("run-store", "data"),
    Output("import-poll-interval", "disabled"),
    Output("start-import", "disabled"),
    Output("fetch-options", "disabled"),
    Output("stop-import", "disabled"),
    Output("run-dropdown", "options"),
    Output("run-dropdown", "value"),
    Output("run-action-message", "children"),
    Output("analysing-run-dropdown", "options"),
    Output("summary-run-dropdown", "options"),
    Input("start-import", "n_clicks"),
    Input("stop-import", "n_clicks"),
    Input("import-poll-interval", "n_intervals"),
    Input("run-dropdown", "value"),
    Input("delete-run-btn", "n_clicks"),
    Input("regen-run-btn", "n_clicks"),
    State("input-file-dropdown", "value"),
    State("year-windows", "value"),
    State("fetch-options", "value"),
    State("run-store", "data"),
    prevent_initial_call=True,
)
def handle_run_actions(
    start_clicks: int,
    stop_clicks: int,
    _poll_count: int,
    selected_run_dir: Optional[str],
    delete_clicks: Optional[int],
    regen_clicks: Optional[int],
    selected_input_file: Optional[str],
    year_windows_value: str,
    fetch_options: List[str],
    run_store: Dict[str, Any],
):
    triggered = dash.ctx.triggered_id
    current_store = dict(run_store or {})
    dropdown_options = no_update
    dropdown_value = no_update
    action_message = no_update

    if triggered == "start-import":
        try:
            windows = _parse_year_windows(year_windows_value)
        except ValueError as exc:
            has_file_now = _input_file_exists(selected_input_file)
            return (
                f"⚠️ {exc}",
                no_update,
                current_store,
                True,
                not has_file_now,
                False,
                True,
                dropdown_options,
                dropdown_value,
                DEFAULT_RUN_MESSAGE if action_message is no_update else action_message,
                no_update, no_update,
            )
        if not selected_input_file:
            return (
                "⚠️ Select an input workbook before importing.",
                no_update,
                current_store,
                True,
                True,
                False,
                True,
                dropdown_options,
                dropdown_value,
                DEFAULT_RUN_MESSAGE if action_message is no_update else action_message,
                no_update, no_update,
            )

        fetch_scopus = "scopus" in (fetch_options or [])
        fetch_unige = "unige" in (fetch_options or [])
        fetch_iris = "iris" in (fetch_options or [])
        try:
            IMPORT_MANAGER.start(
                input_workbook=selected_input_file,
                year_windows=windows,
                fetch_scopus=fetch_scopus,
                fetch_unige=fetch_unige,
                fetch_iris=fetch_iris,
            )
        except RuntimeError as exc:
            state = IMPORT_MANAGER.get_state()
            is_running = state.get("status") == "running"
            has_file_now = _input_file_exists(selected_input_file)
            return (
                f"⚠️ {exc}",
                "\n".join(state.get("logs", [])),
                current_store,
                state.get("status") != "running",
                is_running or not has_file_now,
                is_running,
                not is_running,
                dropdown_options,
                dropdown_value,
                DEFAULT_RUN_MESSAGE if action_message is no_update else action_message,
                no_update, no_update,
            )

    if triggered == "run-dropdown":
        dropdown_options = _run_dropdown_options()
        values = {option["value"] for option in dropdown_options}
        dropdown_value = selected_run_dir if selected_run_dir in values else dropdown_options[0]["value"] if dropdown_options else None
        current_store = _load_run_store_for_value(dropdown_value)
        action_message = (
            f"📂 Viewing run {Path(dropdown_value).name}" if dropdown_value else DEFAULT_RUN_MESSAGE
        )
    if triggered == "stop-import":
        IMPORT_MANAGER.stop()
    elif triggered == "delete-run-btn":
        if selected_run_dir and _delete_run_directory(selected_run_dir):
            action_message = f"🗑️ Deleted run {Path(selected_run_dir).name}."
        else:
            action_message = "⚠️ Unable to delete run."
        dropdown_options, dropdown_value = _sync_run_dropdown(None)
        current_store = _load_run_store_for_value(dropdown_value)
    elif triggered == "regen-run-btn":
        action_message = _regenerate_run_outputs(selected_run_dir)
        dropdown_options, dropdown_value = _sync_run_dropdown(selected_run_dir)
        current_store = _load_run_store_for_value(dropdown_value)

    state = IMPORT_MANAGER.get_state()
    interval_disabled = state.get("status") != "running"
    run_data = current_store
    result = state.get("result")
    if result and state.get("status") in {"completed", "cancelled"}:
        run_data = result
        if triggered == "start-import":
            dropdown_options, dropdown_value = _sync_run_dropdown(run_data.get("run_dir"))
            action_message = "✅ Import completed and outputs generated."

    is_running = state.get("status") == "running"
    has_file = _input_file_exists(selected_input_file)
    start_disabled = is_running or not has_file

    final_message = (
        action_message
        if action_message is not no_update
        else (DEFAULT_RUN_MESSAGE if dropdown_value in (None, no_update) else no_update)
    )

    if dropdown_options is no_update:
        analysing_opts = no_update
        summary_opts   = no_update
    else:
        analysing_opts = [{"label": "All runs", "value": "__all__"}] + dropdown_options
        summary_opts   = dropdown_options

    return (
        _format_import_status(state),
        "\n".join(state.get("logs") or []),
        run_data,
        interval_disabled,
        start_disabled,
        is_running,
        not is_running,
        dropdown_options,
        dropdown_value,
        final_message,
        analysing_opts,
        summary_opts,
    )


@app.callback(
    Output("current-run-label", "children"),
    Output("member-table", "data"),
    Output("member-table", "style_data_conditional"),
    Input("run-store", "data"),
    Input("member-search", "value"),
)
def update_run_view(run_store: Dict[str, Any], search: Optional[str]):
    run_data = run_store or {}
    payloads = run_data.get("payloads") or []
    q = (search or "").strip().lower()
    rows: List[Dict[str, Any]] = []
    for i, payload in enumerate(payloads):
        surname = payload.get("surname", "")
        name    = payload.get("name", "")
        ssd      = payload.get("ssd", "")
        ssd_name = payload.get("ssd_name", "")
        if q and not any(q in s.lower() for s in [surname, name, ssd, ssd_name]):
            continue
        rows.append({
            "inspect":       "🔍",
            "surname":       surname,
            "name":          name,
            "ssd":           f"{ssd} {ssd_name}".strip() if ssd_name else ssd,
            "role":          payload.get("grade") or payload.get("role", ""),
            "_payload_idx":  i,
        })
    metadata = run_data.get("metadata") or {}
    input_file = metadata.get("input_file", "")
    count = metadata.get("source_count", len(payloads))
    label = f"{input_file}  –  {count} members" if input_file else (f"{count} members" if payloads else "No data")
    return label, rows, _TABLE_STYLE_BASE


@app.callback(
    Output("download-summary-btn", "disabled"),
    Input("run-store", "data"),
)
def handle_download_button_state(run_store: Dict[str, Any]) -> bool:
    metadata = (run_store or {}).get("metadata") or {}
    summary_path = metadata.get("summary_path")
    if not summary_path:
        return True
    return not Path(summary_path).exists()


@app.callback(
    Output("download-summary", "data"),
    Input("download-summary-btn", "n_clicks"),
    State("run-store", "data"),
    prevent_initial_call=True,
)
def trigger_summary_download(n_clicks: int, run_store: Dict[str, Any]):
    if not n_clicks:
        return dash.no_update
    metadata = (run_store or {}).get("metadata") or {}
    summary_path = metadata.get("summary_path")
    if not summary_path:
        return dash.no_update
    path = Path(summary_path)
    if not path.exists():
        return dash.no_update
    return dcc.send_file(path, filename=path.name)


@app.callback(
    Output("member-detail", "children"),
    Output("member-table", "style_data_conditional", allow_duplicate=True),
    Output("selected-member-idx", "data"),
    Output("member-pdf-btn", "disabled"),
    Input("member-table", "active_cell"),
    State("run-store", "data"),
    State("member-table", "data"),
    prevent_initial_call=True,
)
def show_member_detail(
    active_cell: Optional[Dict[str, Any]],
    run_store: Dict[str, Any],
    table_data: Optional[List[Dict[str, Any]]],
):
    if not active_cell or active_cell.get("column_id") != "inspect":
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update
    payloads = (run_store or {}).get("payloads") or []
    table_row = active_cell.get("row")
    if table_row is None or not table_data or table_row >= len(table_data):
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update
    payload_idx = table_data[table_row].get("_payload_idx", table_row)
    if payload_idx < 0 or payload_idx >= len(payloads):
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update
    return (
        _member_detail_component(payloads[payload_idx]),
        _table_style_with_row(table_row),
        payload_idx,
        False,
    )


def _make_result_table(result_df) -> dash_table.DataTable:
    columns = [{"name": col, "id": col} for col in result_df.columns]
    data = result_df.fillna("").astype(str).to_dict("records")
    return dash_table.DataTable(
        columns=columns,
        data=data,
        style_as_list_view=True,
        style_table={"overflowX": "auto"},
        style_header={
            "backgroundColor": "#f8f9fa",
            "fontWeight": "600",
            "fontSize": 13,
            "color": "#495057",
            "borderBottom": "2px solid #dee2e6",
            "borderTop": "none",
            "padding": "8px 10px",
        },
        style_cell={
            "textAlign": "left",
            "padding": "8px 10px",
            "fontSize": 13,
            "color": "#212529",
            "borderBottom": "1px solid #f0f0f0",
            "fontFamily": "inherit",
        },
        style_data_conditional=_TABLE_STYLE_BASE,
        sort_action="native",
        sort_mode="single",
        page_action="none",
        cell_selectable=False,
    )


def _history_panel(history: List[Dict[str, Any]]) -> html.Div:
    if not history:
        return html.Div("No questions asked yet.", className="text-muted")
    items = []
    for entry in reversed(history):
        ts    = entry.get("timestamp", "")
        q     = entry.get("question", "")
        code  = entry.get("code", "")
        n     = entry.get("n_rows", 0)
        result_json = entry.get("result_json")

        code_block = html.Details(
            [
                html.Summary("Generated code", style={"cursor": "pointer", "color": "#6c757d", "fontSize": "0.8rem"}),
                html.Pre(code, style={"fontSize": "0.75rem", "backgroundColor": "#f8f9fa",
                                      "padding": "8px", "borderRadius": "4px", "overflowX": "auto", "marginTop": "4px"}),
            ]
        )
        result_block = html.Details(
            [
                html.Summary("Results", style={"cursor": "pointer", "color": "#6c757d", "fontSize": "0.8rem"}),
                html.Div(
                    _make_result_table(__import__("pandas").read_json(result_json, orient="records"))
                    if result_json else html.Span("—", className="text-muted"),
                    style={"marginTop": "4px"},
                ),
            ]
        ) if result_json else html.Div()

        items.append(html.Div(
            [
                html.Div(
                    [html.Span(f"{ts} · ", className="text-muted"), html.Strong(q)],
                    className="mb-1",
                    style={"fontSize": "0.85rem"},
                ),
                html.Div(f"{n} row(s)", className="text-muted", style={"fontSize": "0.75rem"}),
                code_block,
                result_block,
            ],
            style={"borderLeft": "3px solid #dee2e6", "paddingLeft": "10px", "marginBottom": "16px"},
        ))
    return html.Div(items)


@app.callback(
    Output("analysis-result", "children"),
    Output("analysis-status", "children"),
    Output("analysis-history", "data"),
    Output("analysis-result-store", "data"),
    Output("analysis-download-btn", "disabled"),
    Input("analysis-ask-btn", "n_clicks"),
    State("analysis-question", "value"),
    State("analysis-model", "value"),
    State("analysing-run-dropdown", "value"),
    State("analysis-history", "data"),
    prevent_initial_call=True,
)
def run_analysis(
    n_clicks: Optional[int],
    question: Optional[str],
    model: Optional[str],
    run_value: Optional[str],
    history: Optional[List],
):
    _no = dash.no_update
    if not n_clicks:
        return _no, _no, _no, _no, _no

    question = (question or "").strip()
    if not question:
        return html.Div("Please enter a question.", className="text-muted"), "", _no, _no, True

    model = (model or _DEFAULT_OLLAMA_MODEL).strip() or _DEFAULT_OLLAMA_MODEL

    ollama_err = _ensure_ollama()
    if ollama_err:
        return dbc.Alert(f"⚠️ {ollama_err}", color="danger"), "Ollama unavailable.", _no, _no, True

    try:
        if not run_value or run_value == "__all__":
            df, records = load_all_runs(SETTINGS.data_dir)
            scope_label = "all runs"
        else:
            run_store = _load_run_store_for_value(run_value)
            payloads  = run_store.get("payloads") or []
            run_name  = Path(run_value).name
            parts     = run_name.split("_")
            run_label = (
                f"{parts[0]}/{parts[1]}/{parts[2]} #{parts[3]}"
                if len(parts) == 4 else run_name
            )
            date_str  = "_".join(parts[:3]) if len(parts) >= 3 else ""
            run_index = int(parts[3]) if len(parts) == 4 else 0
            df, records = df_from_payloads(payloads, run_label=run_label,
                                           run_date=date_str, run_index=run_index)
            scope_label = f"run {run_label}"
    except Exception as exc:
        return dbc.Alert(f"Failed to load run data: {exc}", color="danger"), "Error loading data.", _no, _no, True

    if df.empty:
        return dbc.Alert("No run data found. Run an import first.", color="warning"), "", _no, _no, True

    try:
        result_df, code = query_llm(
            question,
            df,
            records,
            ollama_url=_DEFAULT_OLLAMA_URL,
            model=model,
        )
    except Exception as exc:
        return dbc.Alert(str(exc), color="danger"), "LLM query failed.", _no, _no, True

    if result_df.empty:
        return dbc.Alert("The query returned no rows.", color="warning"), f"0 rows — {scope_label} — model: {model}", _no, _no, True

    table = _make_result_table(result_df)
    code_block = html.Details(
        [
            html.Summary("Generated code", style={"cursor": "pointer", "color": "#6c757d", "userSelect": "none"}),
            html.Pre(code, style={"fontSize": "0.75rem", "backgroundColor": "#f8f9fa",
                                  "padding": "8px", "borderRadius": "4px", "overflowX": "auto", "marginTop": "6px"}),
        ],
        style={"marginTop": "12px"},
    )

    result_json = result_df.to_json(orient="records")
    from datetime import datetime as _dt
    ts = _dt.now().strftime("%Y-%m-%d %H:%M:%S")
    new_entry = {"timestamp": ts, "question": question, "code": code,
                 "n_rows": len(result_df), "result_json": result_json}
    # Merge in-memory history with whatever is on disk (handles page-refresh gaps)
    disk_history = _load_history(run_value)
    base = history if history is not None else disk_history
    updated_history = list(base) + [new_entry]
    _save_history(run_value, updated_history)

    result_store = {"result_json": result_json, "filename": "analysis_result.xlsx"}

    return (
        html.Div([code_block, table]),
        f"{len(result_df)} row(s) — {scope_label} — model: {model}",
        updated_history,
        result_store,
        False,
    )


@app.callback(
    Output("member-pdf-download", "data"),
    Input("member-pdf-btn", "n_clicks"),
    State("selected-member-idx", "data"),
    State("run-store", "data"),
    prevent_initial_call=True,
)
def download_member_pdf(
    n_clicks: Optional[int],
    member_idx: Optional[int],
    run_store: Dict[str, Any],
):
    if not n_clicks or member_idx is None:
        return dash.no_update
    payloads = (run_store or {}).get("payloads") or []
    if member_idx < 0 or member_idx >= len(payloads):
        return dash.no_update
    payload = payloads[member_idx]
    try:
        pdf_bytes = generate_member_pdf(payload)
    except Exception:
        return dash.no_update
    surname  = re.sub(r"[^A-Za-z0-9_-]", "_", payload.get("surname", "member"))
    name_    = re.sub(r"[^A-Za-z0-9_-]", "_", payload.get("name", ""))
    filename = f"{surname}_{name_}_profile.pdf".strip("_")
    import base64 as _b64
    encoded = _b64.b64encode(pdf_bytes).decode("ascii")
    return {"base64": True, "content": encoded, "filename": filename, "type": "application/pdf"}


@app.callback(
    Output("analysis-history", "data", allow_duplicate=True),
    Input("analysing-run-dropdown", "value"),
    prevent_initial_call="initial_duplicate",
)
def load_history_for_run(run_value: Optional[str]):
    return _load_history(run_value)


@app.callback(
    Output("analysis-history-panel", "children"),
    Input("analysis-history", "data"),
)
def update_history_panel(history: Optional[List]):
    return _history_panel(history or [])


@app.callback(
    Output("analysis-history", "data", allow_duplicate=True),
    Input("analysis-clear-history-btn", "n_clicks"),
    State("analysing-run-dropdown", "value"),
    prevent_initial_call=True,
)
def clear_history(n_clicks: Optional[int], run_value: Optional[str]):
    if not n_clicks:
        return dash.no_update
    hf = _history_file(run_value)
    if hf and hf.exists():
        try:
            hf.unlink()
        except Exception:
            pass
    return []


@app.callback(
    Output("analysis-download", "data"),
    Input("analysis-download-btn", "n_clicks"),
    State("analysis-result-store", "data"),
    prevent_initial_call=True,
)
def download_analysis(n_clicks: Optional[int], result_store: Optional[Dict[str, Any]]):
    if not n_clicks or not result_store:
        return dash.no_update
    result_json = result_store.get("result_json")
    if not result_json:
        return dash.no_update
    import pandas as pd_
    result_df = pd_.read_json(result_json, orient="records")
    return dcc.send_data_frame(result_df.to_excel, "analysis_result.xlsx", index=False)


def _summary_score_badge(val: Optional[float]) -> html.Span:
    if val is None:
        return html.Span("—", className="text-muted")
    if val >= 1.2:  color = "success"
    elif val >= 0.8: color = "primary"
    elif val >= 0.4: color = "warning"
    else:            color = "danger"
    return dbc.Badge(f"{val:.2f}", color=color, className="me-1")


def _metrics_from_payload(p: Dict[str, Any]) -> Dict[str, Optional[float]]:
    """Extract flat numeric metrics from one payload."""
    mbs: Dict[str, Dict[str, Any]] = {}
    for m in p.get("scopus_metrics") or []:
        period = (m.get("period") or "").strip()
        for prefix, sfx in [("05 years", "5y"), ("10 years", "10y"), ("15 years", "15y")]:
            if period.startswith(prefix):
                mbs[sfx] = m
        if period.lower() == "absolute":
            mbs["abs"] = m

    def _m(sfx: str, key: str) -> Optional[float]:
        v = mbs.get(sfx, {}).get(key)
        return float(v) if v is not None else None

    live_scores = compute_scores(p.get("ssd"), p.get("scopus_metrics") or [], _THRESHOLDS)
    def _s(ind: str) -> Optional[float]:
        v = (live_scores.get(ind) or {}).get("score")
        return float(v) if v is not None else None

    return {
        "h_5y":   _m("5y",  "hindex"),  "h_10y":  _m("10y", "hindex"),
        "h_15y":  _m("15y", "hindex"),  "h_abs":  _m("abs", "hindex"),
        "c_5y":   _m("5y",  "citations"), "c_10y": _m("10y", "citations"),
        "c_15y":  _m("15y", "citations"), "c_abs": _m("abs", "citations"),
        "p_5y":   _m("5y",  "total_products"), "p_10y": _m("10y", "total_products"),
        "p_15y":  _m("15y", "total_products"), "p_abs": _m("abs", "total_products"),
        "s_art":  _s("articles"),
        "s_cit":  _s("citations"),
        "s_h":    _s("hindex"),
    }


def _avgs(rows: List[Dict[str, Optional[float]]], key: str) -> Optional[float]:
    vals = [r[key] for r in rows if r.get(key) is not None]
    return sum(vals) / len(vals) if vals else None


def _fmt(v: Optional[float], decimals: int = 1) -> str:
    return f"{v:.{decimals}f}" if v is not None else "—"


def _dept_metrics_table(rows: List[Dict[str, Optional[float]]]) -> html.Table:
    _TH = {"backgroundColor": "#f8f9fa", "fontWeight": "600", "fontSize": 12,
           "color": "#495057", "padding": "6px 10px", "borderBottom": "2px solid #dee2e6",
           "textAlign": "center", "whiteSpace": "nowrap"}
    _TD = {"padding": "6px 10px", "fontSize": 13, "borderBottom": "1px solid #f0f0f0",
           "textAlign": "center", "color": "#212529"}
    _TDL = {**_TD, "textAlign": "left", "fontWeight": "600", "color": "#495057"}

    def _score_cell(key: str) -> html.Td:
        return html.Td(_summary_score_badge(_avgs(rows, key)), style=_TD)

    def _num_cell(key: str, dec: int = 1) -> html.Td:
        return html.Td(_fmt(_avgs(rows, key), dec), style=_TD)

    header = html.Thead(html.Tr([
        html.Th("", style=_TH),
        html.Th("5y",  style=_TH), html.Th("10y", style=_TH),
        html.Th("15y", style=_TH), html.Th("Abs", style=_TH),
        html.Th("Score", style={**_TH, "borderLeft": "2px solid #dee2e6"}),
    ]))
    body = html.Tbody([
        html.Tr([
            html.Td("H-index",   style=_TDL),
            _num_cell("h_5y", 1), _num_cell("h_10y", 1),
            _num_cell("h_15y", 1), _num_cell("h_abs", 1),
            html.Td(_summary_score_badge(_avgs(rows, "s_h")),
                    style={**_TD, "borderLeft": "2px solid #dee2e6"}),
        ], style={"backgroundColor": "#fafafa"}),
        html.Tr([
            html.Td("Citations", style=_TDL),
            _num_cell("c_5y", 0), _num_cell("c_10y", 0),
            _num_cell("c_15y", 0), _num_cell("c_abs", 0),
            html.Td(_summary_score_badge(_avgs(rows, "s_cit")),
                    style={**_TD, "borderLeft": "2px solid #dee2e6"}),
        ]),
        html.Tr([
            html.Td("Products",  style=_TDL),
            _num_cell("p_5y", 1), _num_cell("p_10y", 1),
            _num_cell("p_15y", 1), _num_cell("p_abs", 1),
            html.Td(_summary_score_badge(_avgs(rows, "s_art")),
                    style={**_TD, "borderLeft": "2px solid #dee2e6"}),
        ], style={"backgroundColor": "#fafafa"}),
    ])
    return html.Table([header, body], className="w-100",
                      style={"borderCollapse": "collapse", "fontSize": 13})


def _ssd_breakdown_table(ssd_metrics: Dict[str, List[Dict]]) -> html.Table:
    _TH = {"backgroundColor": "#f8f9fa", "fontWeight": "600", "fontSize": 11,
           "color": "#495057", "padding": "5px 8px", "borderBottom": "2px solid #dee2e6",
           "textAlign": "center", "whiteSpace": "nowrap"}
    _TD = {"padding": "5px 8px", "fontSize": 12, "borderBottom": "1px solid #f0f0f0",
           "textAlign": "center", "color": "#212529"}
    _TDL = {**_TD, "textAlign": "left", "fontWeight": "600", "color": "#495057",
            "maxWidth": "140px", "overflow": "hidden", "textOverflow": "ellipsis",
            "whiteSpace": "nowrap"}

    header = html.Thead(html.Tr([
        html.Th("SSD",        style={**_TH, "textAlign": "left"}),
        html.Th("N",          style=_TH),
        html.Th("H 5y",       style=_TH), html.Th("H 10y",   style=_TH), html.Th("H 15y",   style=_TH),
        html.Th("Cit 5y",     style=_TH), html.Th("Cit 10y", style=_TH), html.Th("Cit 15y", style=_TH),
        html.Th("Prod 5y",    style=_TH), html.Th("Prod 10y",style=_TH), html.Th("Prod 15y",style=_TH),
        html.Th("Sc. Art.",   style={**_TH, "borderLeft": "2px solid #dee2e6"}),
        html.Th("Sc. Cit.",   style=_TH),
        html.Th("Sc. H",      style=_TH),
    ]))

    def _score_avg(rows: List[Dict]) -> float:
        vals = [v for k in ("s_art", "s_cit", "s_h") for v in [_avgs(rows, k)] if v is not None]
        return sum(vals) / len(vals) if vals else -1.0

    body_rows = []
    for i, (ssd, rows) in enumerate(
        sorted(ssd_metrics.items(), key=lambda x: -_score_avg(x[1]))
    ):
        bg = "#fafafa" if i % 2 == 0 else "#ffffff"
        body_rows.append(html.Tr([
            html.Td(ssd,       style={**_TDL, "backgroundColor": bg}),
            html.Td(len(rows), style={**_TD,  "backgroundColor": bg, "fontWeight": "600"}),
            html.Td(_fmt(_avgs(rows, "h_5y"),   1), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "h_10y"),  1), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "h_15y"),  1), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "c_5y"),   0), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "c_10y"),  0), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "c_15y"),  0), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "p_5y"),   1), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "p_10y"),  1), style={**_TD, "backgroundColor": bg}),
            html.Td(_fmt(_avgs(rows, "p_15y"),  1), style={**_TD, "backgroundColor": bg}),
            html.Td(_summary_score_badge(_avgs(rows, "s_art")),
                    style={**_TD, "backgroundColor": bg, "borderLeft": "2px solid #dee2e6"}),
            html.Td(_summary_score_badge(_avgs(rows, "s_cit")),
                    style={**_TD, "backgroundColor": bg}),
            html.Td(_summary_score_badge(_avgs(rows, "s_h")),
                    style={**_TD, "backgroundColor": bg}),
        ]))

    return html.Table([header, html.Tbody(body_rows)], className="w-100",
                      style={"borderCollapse": "collapse", "fontSize": 12})


def _build_ssd_export_data(ssd_metrics: Dict[str, List[Dict]]) -> Optional[str]:
    """Serialise the SSD breakdown as a JSON string for the export store."""
    def _score_avg(rows):
        vals = [v for k in ("s_art", "s_cit", "s_h") for v in [_avgs(rows, k)] if v is not None]
        return round(sum(vals) / len(vals), 3) if vals else None

    records = []
    for ssd, rows in sorted(ssd_metrics.items(), key=lambda x: -(_score_avg(x[1]) or -1)):
        records.append({
            "SSD": ssd,
            "N": len(rows),
            "H 5y":    _avgs(rows, "h_5y"),
            "H 10y":   _avgs(rows, "h_10y"),
            "H 15y":   _avgs(rows, "h_15y"),
            "Cit 5y":  _avgs(rows, "c_5y"),
            "Cit 10y": _avgs(rows, "c_10y"),
            "Cit 15y": _avgs(rows, "c_15y"),
            "Prod 5y":  _avgs(rows, "p_5y"),
            "Prod 10y": _avgs(rows, "p_10y"),
            "Prod 15y": _avgs(rows, "p_15y"),
            "Score articles":  _avgs(rows, "s_art"),
            "Score citations": _avgs(rows, "s_cit"),
            "Score h-index":   _avgs(rows, "s_h"),
        })
    return json.dumps(records)


@app.callback(
    Output("summary-content", "children"),
    Output("ssd-export-store", "data"),
    Input("summary-run-dropdown", "value"),
)
def update_summary(selected_run: Optional[str]):
    _no = dash.no_update
    if not selected_run:
        return html.Div("Select a run above to see the department summary.", className="text-muted"), _no
    run_data = _load_run_store_for_value(selected_run)
    payloads = run_data.get("payloads") or []
    if not payloads:
        return html.Div("No data found for the selected run.", className="text-muted"), _no

    all_rows: List[Dict[str, Optional[float]]] = []
    ssd_metrics: Dict[str, List[Dict]] = {}

    for p in payloads:
        m = _metrics_from_payload(p)
        all_rows.append(m)
        ssd      = p.get("ssd") or "Unknown"
        ssd_name = p.get("ssd_name", "")
        ssd_key  = f"{ssd} {ssd_name}".strip() if ssd_name else ssd
        ssd_metrics.setdefault(ssd_key, []).append(m)

    total    = len(payloads)
    n_ssds   = len(ssd_metrics)
    run_dir  = run_data.get("run_dir") or ""
    run_name = Path(run_dir).name if run_dir else "—"

    return html.Div([
        # ── KPI strip ────────────────────────────────────────────────────────
        dbc.Row([
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div("Members", className="text-muted small mb-1"),
                html.H3(total, className="mb-0"),
            ]), className="text-center shadow-sm"), md=2),
            dbc.Col(dbc.Card(dbc.CardBody([
                html.Div("SSDs", className="text-muted small mb-1"),
                html.H3(n_ssds, className="mb-0"),
            ]), className="text-center shadow-sm"), md=2),
        ], className="g-3 mb-4"),

        # ── Department averages ───────────────────────────────────────────────
        dbc.Card(dbc.CardBody([
            html.H5("Department averages", className="mb-3"),
            html.P(
                "Score column shows the average threshold score (D.M. 589/2018): "
                "≥1.2 Evaluator · ≥0.8 Full Prof. · ≥0.4 Assoc. Prof.",
                className="text-muted small mb-3",
            ),
            _dept_metrics_table(all_rows),
        ]), className="shadow-sm mb-4"),

        # ── By SSD ───────────────────────────────────────────────────────────
        dbc.Card(dbc.CardBody([
            dbc.Row([
                dbc.Col(html.H5("SSD", className="mb-0"), className="align-self-center"),
                dbc.Col(
                    dbc.Button("Export Excel", id="ssd-export-btn", color="success",
                               size="sm", disabled=False),
                    width="auto",
                ),
            ], className="g-2 mb-3 align-items-center"),
            html.Div(
                _ssd_breakdown_table(ssd_metrics),
                style={"overflowX": "auto"},
            ),
        ]), className="shadow-sm mb-3"),

        html.Div(f"Run: {run_name}", className="text-muted small"),
    ]), _build_ssd_export_data(ssd_metrics)


@app.callback(
    Output("compare-members-dropdown", "options"),
    Input("run-store", "data"),
)
def populate_compare_dropdown(run_store: Dict[str, Any]):
    payloads = (run_store or {}).get("payloads") or []
    return [
        {"label": f"{p.get('surname', '')} {p.get('name', '')}", "value": i}
        for i, p in enumerate(payloads)
    ]


@app.callback(
    Output("comparison-result", "children"),
    Output("comparison-export-store", "data"),
    Output("comparison-export-btn", "disabled"),
    Input("compare-btn", "n_clicks"),
    Input("compare-members-dropdown", "value"),
    State("run-store", "data"),
    prevent_initial_call=True,
)
def compare_members(n_clicks: Optional[int], selected: Optional[List[int]], run_store: Dict[str, Any]):
    _no = dash.no_update
    if not selected:
        return html.Div(), None, True
    payloads = (run_store or {}).get("payloads") or []
    chosen = [payloads[i] for i in selected if 0 <= i < len(payloads)]
    if not chosen:
        return html.Div("No members selected.", className="text-muted"), None, True

    live_scores_list = [
        compute_scores(p.get("ssd"), p.get("scopus_metrics") or [], _THRESHOLDS)
        for p in chosen
    ]

    _TH  = {"padding": "6px 10px", "fontSize": 12, "fontWeight": "600",
             "backgroundColor": "#f8f9fa", "color": "#495057",
             "borderBottom": "2px solid #dee2e6", "whiteSpace": "nowrap"}
    _TD  = {"padding": "5px 10px", "fontSize": 12, "borderBottom": "1px solid #f0f0f0",
             "textAlign": "center", "color": "#212529"}
    _TDL = {**_TD, "textAlign": "left", "color": "#6c757d", "fontWeight": "600"}
    _TDG = {**_TDL, "backgroundColor": "#f8f9fa", "fontSize": 11,
             "color": "#0d6efd", "paddingTop": "8px"}

    def _metric_val(p, field, period_prefix):
        for m in p.get("scopus_metrics") or []:
            if (m.get("period") or "").startswith(period_prefix):
                v = m.get(field)
                return str(v) if v is not None else "—"
        return "—"

    def _score_badge(score):
        if score is None:
            return html.Span("—", className="text-muted")
        if score >= 1.2:   color = "success"
        elif score >= 0.8: color = "primary"
        elif score >= 0.4: color = "warning"
        else:              color = "danger"
        return dbc.Badge(f"{score:.1f}", color=color)

    def _ratio_cell(block, level_key):
        lvl = (block or {}).get(level_key) or {}
        v, t, r = lvl.get("value"), lvl.get("threshold"), lvl.get("ratio")
        if v is None or t is None:
            return html.Td("—", style=_TD)
        text = f"{v}/{t}"
        ratio_color = "#198754" if r and r >= 1 else "#dc3545"
        ratio_span = html.Span(f" ={r:.2f}" if r is not None else "", style={"color": ratio_color, "fontWeight": "600"})
        return html.Td([text, ratio_span], style=_TD)

    def _group_header(label):
        return html.Tr(
            html.Td(label, colSpan=1 + len(chosen), style=_TDG)
        )

    def _simple_row(label, cells_content, bg=None):
        style = {**_TDL, **({"backgroundColor": bg} if bg else {})}
        return html.Tr(
            [html.Td(label, style=style)] + cells_content,
            style={"borderBottom": "1px solid #f0f0f0"},
        )

    headers = html.Tr(
        [html.Th("", style=_TH)] +
        [html.Th(f"{p.get('surname','')} {p.get('name','')}", style={**_TH, "textAlign": "center"})
         for p in chosen]
    )

    rows = []

    # ── Bibliometric metrics ──────────────────────────────────────────────────
    rows.append(_group_header("Bibliometric metrics"))
    for label, field, period in [
        ("H-index 5y",    "hindex",         "05 years"),
        ("H-index 10y",   "hindex",         "10 years"),
        ("H-index 15y",   "hindex",         "15 years"),
        ("Citations 10y", "citations",      "10 years"),
        ("Citations 15y", "citations",      "15 years"),
        ("Products 5y",   "total_products", "05 years"),
        ("Products 10y",  "total_products", "10 years"),
    ]:
        cells = [html.Td(_metric_val(p, field, period), style=_TD) for p in chosen]
        rows.append(_simple_row(label, cells))

    # ── Score + ratios per indicator ──────────────────────────────────────────
    for ind_key, ind_label in [("articles", "Articles"), ("citations", "Citations"), ("hindex", "H-index")]:
        rows.append(_group_header(f"Threshold scores — {ind_label}"))

        # Score row
        score_cells = [
            html.Td(_score_badge((ls.get(ind_key) or {}).get("score")), style=_TD)
            for ls in live_scores_list
        ]
        rows.append(_simple_row("Score", score_cells, bg="#fff8f0"))

        # Ratio rows per level
        for level_key, level_label in [
            ("ii_fascia",    "Assoc. Prof."),
            ("i_fascia",     "Full Prof."),
            ("commissario",  "Evaluator"),
        ]:
            ratio_cells = [_ratio_cell(ls.get(ind_key), level_key) for ls in live_scores_list]
            rows.append(_simple_row(level_label, ratio_cells))

    # Build flat records for Excel export
    export_records = []
    for p, ls in zip(chosen, live_scores_list):
        name = f"{p.get('surname','')} {p.get('name','')}".strip()
        rec: Dict[str, Any] = {"Member": name, "SSD": p.get("ssd", "")}
        for field, period in [
            ("H-index 5y",    ("hindex",         "05 years")),
            ("H-index 10y",   ("hindex",         "10 years")),
            ("H-index 15y",   ("hindex",         "15 years")),
            ("Citations 10y", ("citations",      "10 years")),
            ("Citations 15y", ("citations",      "15 years")),
            ("Products 5y",   ("total_products", "05 years")),
            ("Products 10y",  ("total_products", "10 years")),
        ]:
            met, pfx = period
            rec[field] = next(
                (m.get(met) for m in (p.get("scopus_metrics") or [])
                 if (m.get("period") or "").startswith(pfx)),
                None,
            )
        for ind in ["articles", "citations", "hindex"]:
            block = ls.get(ind) or {}
            rec[f"Score {ind}"] = block.get("score")
            for lk, ll in [("ii_fascia", "Assoc.Prof"), ("i_fascia", "Full Prof"), ("commissario", "Evaluator")]:
                lvl = block.get(lk) or {}
                rec[f"{ind.capitalize()} {ll} ratio"] = lvl.get("ratio")
        export_records.append(rec)

    return (
        html.Div(
            html.Table(
                [html.Thead(headers), html.Tbody(rows)],
                className="w-100",
                style={"borderCollapse": "collapse", "fontSize": "0.85rem"},
            ),
            style={"overflowX": "auto"},
        ),
        json.dumps(export_records),
        False,
    )


@app.callback(
    Output("ssd-export-download", "data"),
    Input("ssd-export-btn", "n_clicks"),
    State("ssd-export-store", "data"),
    prevent_initial_call=True,
)
def download_ssd_export(n_clicks: Optional[int], store_json: Optional[str]):
    if not n_clicks or not store_json:
        return dash.no_update
    import pandas as _pd
    df = _pd.DataFrame(json.loads(store_json))
    return dcc.send_data_frame(df.to_excel, "ssd_summary.xlsx", index=False)


@app.callback(
    Output("comparison-export-download", "data"),
    Input("comparison-export-btn", "n_clicks"),
    State("comparison-export-store", "data"),
    prevent_initial_call=True,
)
def download_comparison_export(n_clicks: Optional[int], store_json: Optional[str]):
    if not n_clicks or not store_json:
        return dash.no_update
    import pandas as _pd
    df = _pd.DataFrame(json.loads(store_json))
    return dcc.send_data_frame(df.to_excel, "member_comparison.xlsx", index=False)


def main() -> None:  # pragma: no cover - manual start
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8050")), debug=False)


if __name__ == "__main__":  # pragma: no cover
    main()

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class DataPreparation:
    def prepare(
        self,
        payloads: Sequence[Dict[str, Any]],
        run_dir: Path,
        input_workbook: str,
    ) -> Path:
        output_dir = run_dir / "elaborations"
        output_dir.mkdir(parents=True, exist_ok=True)
        summary_name = f"{Path(input_workbook).stem}_results.xlsx"
        summary_path = output_dir / summary_name

        records = [self._build_summary_row(payload) for payload in payloads]
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"

        if records:
            header = list(records[0].keys())
            sheet.append(header)
            for record in records:
                sheet.append([record.get(column, "") for column in header])
        else:
            sheet.append(["No data available"])

        self._autosize_columns(sheet)
        workbook.save(summary_path)
        return summary_path

    def _build_summary_row(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        metrics = payload.get("scopus_metrics") or []
        absolute = next(
            (metric for metric in metrics if (metric.get("period") or "").lower() == "absolute"),
            {},
        )

        row = {
            "Surname": payload.get("surname", ""),
            "Name": payload.get("name", ""),
            "Unit": payload.get("unit", ""),
            "Role": payload.get("role", ""),
            "SSD": payload.get("ssd", ""),
            "scopus_id": payload.get("scopus_id", ""),
            "unige_id": payload.get("unige_id", ""),
            "products": absolute.get("total_products", ""),
            "citations": absolute.get("citations", ""),
            "h_index": absolute.get("hindex", ""),
        }

        for metric in metrics:
            period = metric.get("period", "")
            suffix = self._extract_suffix(period)
            if not suffix:
                continue
            label = suffix.replace("y", "")
            row[f"products_{label}"] = metric.get("total_products", "")
            row[f"citations_{label}"] = metric.get("citations", "")
            row[f"journals_{label}"] = metric.get("journals", "")
            row[f"conferences_{label}"] = metric.get("conferences", "")
            row[f"h_index_{label}"] = metric.get("hindex", "")

        return row

    @staticmethod
    def _extract_suffix(period: str) -> str:
        digits = ""
        for char in period:
            if char.isdigit():
                digits += char
            elif digits:
                break
        return f"{digits}y" if digits else ""

    @staticmethod
    def _autosize_columns(sheet: Any) -> None:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

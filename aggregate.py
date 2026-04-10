from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

from member import Member

from openpyxl import load_workbook

# Positional column names used when the workbook has no header row.
# Column order: surname, name, grade/role, ssd, scopus_id [, unige_id, unit]
_POSITIONAL_HEADER = ["surname", "name", "grade", "ssd", "scopusid", "unigeid", "unit"]
_REQUIRED_HEADER_FIELDS = {"surname", "name", "scopusid"}


class Aggregate:
    """Load members from an Excel workbook."""

    def __init__(self, input_workbook: str) -> None:
        self.input_workbook = Path(input_workbook)

    def load_members(self) -> List[Member]:
        if not self.input_workbook.exists():
            raise FileNotFoundError(f"Input workbook not found: {self.input_workbook}")
        suffix = self.input_workbook.suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise ValueError(f"Unsupported roster format '{suffix}'. Provide an XLSX workbook.")

        rows = self._read_rows_from_xlsx()
        members: List[Member] = []
        for row in rows:
            normalized = self._normalize_row(row)

            surname = normalized.get("surname")
            name = normalized.get("name")
            scopus_id = normalized.get("scopusid")
            if not (surname and name and scopus_id):
                continue

            unige_id = normalized.get("unigeid") or None
            unit = normalized.get("unit") or None
            # "role" comes from headered files (column "Role"); "grade" from positional mapping
            grade = self._normalize_grade(normalized.get("grade") or normalized.get("role") or None)
            ssd = normalized.get("ssd") or None

            members.append(
                Member(
                    surname=surname,
                    name=name,
                    scopus_id=scopus_id,
                    unit=unit,
                    unige_id=unige_id,
                    grade=grade,
                    ssd=ssd,
                )
            )
        return members

    def _read_rows_from_xlsx(self) -> List[Dict[Optional[str], Optional[str]]]:
        workbook = load_workbook(filename=self.input_workbook, read_only=True, data_only=True)
        sheet = workbook.active

        all_rows: List[List[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [self._cell_to_text(cell) for cell in row]
            if any(values):
                all_rows.append(values)
        workbook.close()

        if not all_rows:
            return []

        # Detect whether the first row is a real header or data.
        normalized_first = {re.sub(r"[^a-z0-9]", "", v.lower()) for v in all_rows[0] if v}
        if _REQUIRED_HEADER_FIELDS & normalized_first:
            header = all_rows[0]
            data_rows = all_rows[1:]
        else:
            # No header row — use positional mapping.
            n_cols = len(all_rows[0])
            header = _POSITIONAL_HEADER[:n_cols]
            data_rows = all_rows

        records: List[Dict[Optional[str], Optional[str]]] = []
        for row in data_rows:
            if not any(row):
                continue
            record: Dict[Optional[str], Optional[str]] = {}
            for idx, key in enumerate(header):
                if not key:
                    continue
                record[key] = row[idx] if idx < len(row) else ""
            records.append(record)

        return records

    _GRADE_ALIASES: dict = {
        "ordinario": "Professore Ordinario",
        "associato": "Professore Associato",
    }

    @classmethod
    def _normalize_grade(cls, value: Optional[str]) -> Optional[str]:
        if not value:
            return None
        return cls._GRADE_ALIASES.get(value.strip().lower(), value.strip())

    @staticmethod
    def _cell_to_text(value: Optional[object]) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @staticmethod
    def _normalize_row(row: Dict[Optional[str], Optional[str]]) -> Dict[str, str]:
        cleaned: Dict[str, str] = {}
        for key, value in row.items():
            if key is None:
                continue
            normalized_key = re.sub(r"[^a-z0-9]", "", key.strip().lower())
            if not normalized_key:
                continue
            cleaned[normalized_key] = (value or "").strip()
        return cleaned

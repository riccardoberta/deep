"""
Bibliometric threshold scoring based on D.M. 589/2018.

Scoring per indicator (articoli, citazioni, h-index):
  0.0  → below II fascia threshold
  0.4  → meets II fascia threshold
  0.8  → meets I fascia threshold
  1.2  → meets Commissari threshold
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Threshold record: 9 values in order (matches soglie.xlsx column order)
#   [0] II fascia  – Art. 5a
#   [1] II fascia  – Cit. 10a
#   [2] II fascia  – H-idx 10a
#   [3] I fascia   – Art. 10a
#   [4] I fascia   – Cit. 15a
#   [5] I fascia   – H-idx 15a
#   [6] Commissari – Art. 10a
#   [7] Commissari – Cit. 15a
#   [8] Commissari – H-idx 15a
# ---------------------------------------------------------------------------
ThresholdRow = Tuple[
    Optional[int], Optional[int], Optional[int],  # II fascia  (score 0.4)
    Optional[int], Optional[int], Optional[int],  # I fascia   (score 0.8)
    Optional[int], Optional[int], Optional[int],  # Commissari (score 1.2)
]

_DEFAULT_XLSX = Path(__file__).parent / "soglie" / "soglie.xlsx"


def load_thresholds(xlsx_path: Path | str = _DEFAULT_XLSX) -> Dict[str, ThresholdRow]:
    """Return a dict mapping SSD code → 9-tuple of threshold values.

    The workbook must have a sheet named 'Soglie' with columns:
      'SSD codice', 'Art. 5a', 'Cit. 10a', 'H-idx 10a',
      'Art. 10a', 'Cit. 15a', 'H-idx 15a',
      'Art. 10a.1', 'Cit. 15a.1', 'H-idx 15a.1'
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    sheet = wb["Soglie"] if "Soglie" in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    def _col(name: str) -> int:
        return header.index(name)

    idx_code = _col("SSD codice")
    # Order in tuple: I fascia, II fascia, Commissario
    idx_vals = (
        _col("Art. 5a - II"),  _col("Cit. 10a - II"),  _col("H-idx 10a - II"),
        _col("Art. 10a - I"),  _col("Cit. 15a - I"),   _col("H-idx 15a - I"),
        _col("Art. 10a - C"),  _col("Cit. 15a - C"),   _col("H-idx 15a - C"),
    )

    def _v(row: tuple, i: int) -> Optional[int]:
        v = row[i] if i < len(row) else None
        return int(v) if isinstance(v, (int, float)) and v is not None else None

    result: Dict[str, ThresholdRow] = {}
    for row in rows[1:]:
        code = row[idx_code] if idx_code < len(row) else None
        if not code:
            continue
        result[str(code).strip()] = tuple(_v(row, i) for i in idx_vals)  # type: ignore[assignment]
    return result


def _find_threshold(
    thresholds: Dict[str, ThresholdRow], ssd_str: str
) -> Optional[ThresholdRow]:
    """Return the threshold row for *ssd_str*, or None if not found."""
    if not ssd_str:
        return None
    return thresholds.get(ssd_str.strip())


def _get_metric(metrics: List[Dict[str, Any]], period_prefix: str, field: str) -> Optional[int]:
    for m in metrics:
        if str(m.get("period", "")).startswith(period_prefix):
            v = m.get(field)
            return int(v) if v is not None else None
    return None


def _level_entry(
    value: Optional[int], threshold: Optional[int], years: int
) -> Dict[str, Any]:
    if value is None or not threshold:
        return {"years": years, "value": value, "threshold": threshold, "ratio": None}
    return {
        "years": years,
        "value": value,
        "threshold": threshold,
        "ratio": round(value / threshold, 2),
    }


def _indicator_block(
    val_ii: Optional[int], thresh_ii: Optional[int], years_ii: int,
    val_i:  Optional[int], thresh_i:  Optional[int], years_i:  int,
    val_c:  Optional[int], thresh_c:  Optional[int], years_c:  int,
) -> Dict[str, Any]:
    if val_ii is None or thresh_ii is None:
        score = None
    else:
        score = 0.0
        if val_ii >= thresh_ii:
            score = 0.4
        if val_i is not None and thresh_i is not None and val_i >= thresh_i:
            score = 0.8
        if val_c is not None and thresh_c is not None and val_c >= thresh_c:
            score = 1.2

    return {
        "score":       score,
        "ii_fascia":   _level_entry(val_ii, thresh_ii, years_ii),
        "i_fascia":    _level_entry(val_i,  thresh_i,  years_i),
        "commissario": _level_entry(val_c,  thresh_c,  years_c),
    }


def compute_scores(
    ssd_str: Optional[str],
    metrics: List[Dict[str, Any]],
    thresholds: Dict[str, ThresholdRow],
) -> Dict[str, Any]:
    """Return a *scores* dict nested under three indicator keys."""
    def _null_block() -> Dict[str, Any]:
        empty = {"years": None, "value": None, "threshold": None, "ratio": None}
        return {"score": None, "ii_fascia": dict(empty), "i_fascia": dict(empty), "commissario": dict(empty)}

    null: Dict[str, Any] = {
        "articles":  _null_block(),
        "citations": _null_block(),
        "hindex":    _null_block(),
    }

    if not ssd_str or not metrics:
        return null

    thresh = _find_threshold(thresholds, ssd_str)
    if thresh is None:
        return null

    (art5a_ii,  cit10a_ii,  h10a_ii,
     art10a_i,  cit15a_i,  h15a_i,
     art10a_c,  cit15a_c,  h15a_c) = thresh

    art5  = _get_metric(metrics, "05 years", "total_products")
    art10 = _get_metric(metrics, "10 years", "total_products")
    cit10 = _get_metric(metrics, "10 years", "citations")
    cit15 = _get_metric(metrics, "15 years", "citations")
    h10   = _get_metric(metrics, "10 years", "hindex")
    h15   = _get_metric(metrics, "15 years", "hindex")

    return {
        "articles": _indicator_block(
            art5,  art5a_ii,   5,   # II fascia  → score 0.4
            art10, art10a_i,  10,   # I fascia   → score 0.8
            art10, art10a_c,  10,   # Commissario → score 1.2
        ),
        "citations": _indicator_block(
            cit10, cit10a_ii,  10,
            cit15, cit15a_i,   15,
            cit15, cit15a_c,   15,
        ),
        "hindex": _indicator_block(
            h10, h10a_ii,  10,
            h15, h15a_i,   15,
            h15, h15a_c,   15,
        ),
    }
